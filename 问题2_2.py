# -*- coding: utf-8 -*-
"""
Created on Wed Oct 30 03:55:10 2019

@author: Administrator
"""

from openpyxl import load_workbook
excel=load_workbook('问题2.xlsx')
#获取sheet：
table = excel.get_sheet_by_name('Exc 2 DATA')   #通过表名获取  
#获取行数和列数：
rows=table.max_row   #获取行数
cols=table.max_column    #获取列数
#获取单元格值：
#Data=table.cell(row=3,column=1).value     
#print(rows,cols)
data = []

for row in range(2,rows+1):#获得所有数据
    c = []
    for col in range(1,6):
        Data=table.cell(row=row,column=col).value     
        if Data == None: #替换excel内的None
            Data = 0
        c.append(Data)
    c.append(sum(c[2:5])) #替换excel内的公式
    if c[-1] > 100: #为大客户，添加供货期限
        c.append(5)
    elif (c[-1] <=100)&(c[-1]>=50):#为中客户，添加供货期限
        c.append(2)
    else:
        c.append(0)#为小客户，添加供货期限
    #为不同期客户 提供开始时间
    if c[1] == 'Phase I':
        c[1] = 0
    elif c[1] == 'Phase II':
        c[1] = 1
    
    #将cmr 换成 rcm
    swap = c[2]
    c[2] = c[4]
    c[4] = c[3]
    c[3] = swap
    data.append(c)
print(data)
print('数据清洗完毕')



#记录每个月的剩余量
rcm = []

#题目只有10个月的格子
for j in range(10):
    
    
    if j == 0:
        rcm.append([0,0,0])
#        print(rcm)
    else:
        rcm.append([i for i in rcm[-1]])
#        print(rcm)
    #设定月产量
    rcm_0 = [100,500,3000]
    
    for i in data:
        #判断一期二期客户
        if i[1] == 1:
            i[1] = i[1]-1
        elif i[1] == 0:
            #判断客户类别
            if i[-1] == 0:
                rcm_0[0] = rcm_0[0] - i[2]
                rcm_0[1] = rcm_0[1] - i[3]
                rcm_0[2] = rcm_0[2] - i[4]
            else:
                i[-1] = i[-1] - 1
#    print(rcm_0)
    for i in range(3):
        rcm[j][i] = rcm[j][i] + rcm_0[i]
print(rcm)
    

   








