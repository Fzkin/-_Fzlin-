# -*- coding: utf-8 -*-
"""
Created on Tue Oct 29 22:01:29 2019

@author: Administrator
"""

from openpyxl import load_workbook
excel=load_workbook('问题2.xlsx')
#获取sheet：
table = excel.get_sheet_by_name('Exc 1 DATA')   #通过表名获取  
#获取行数和列数：
rows=table.max_row   #获取行数
cols=table.max_column    #获取列数
#获取单元格值：
#Data=table.cell(row=3,column=1).value     
#print(rows,cols)
data = []

for row in range(3,rows+1):#获得所有数据
    c = []
    for col in range(1,6):
        Data=table.cell(row=row,column=col).value     
        c.append(Data)
    data.append(c)
def number(i):#统计所有数据
    count_revenue_large = 0
    count_revenue_med = 0
    count_revenue_small = 0
    count_large = 0
    count_med = 0
    count_small = 0
    for j in data:
        if j[i+1] >= 1000000:
            count_revenue_large += j[i+1]
            if j[0] != None:        #匿名账户不计数，但记税收。
                count_large += 1
            
        elif j[i+1] < 100000:
            count_revenue_small += j[i+1]
            if j[0] != None:
                count_small += 1
            
        elif (j[i+1] < 1000000)&(j[i+1] >= 100000):
            count_revenue_med += j[i+1] 
            if j[0] != None:
                count_med += 1
            
    return [count_revenue_large,count_revenue_med,count_revenue_small,count_large,count_med,count_small]
    

data_1 = []
for i in range(0,4):#获得所有月份总税收和账号等级数量
    a = number(i)
    if i ==0:
        a.append(31)
    elif i ==1:
        a.append(28)
    elif i ==2:
        a.append(31)
    elif i ==3:
        a.append(30)
    data_1.append(a)



#标准化数据
#data_1[0][0] = data_1[0][0] / 31 * 31 / data_1[0][3] 
#data_1[1][0] = data_1[1][0] / 28 * 31 / data_1[1][3] 
#data_1[2][0] = data_1[2][0] / 31 * 31 / data_1[2][3] 
#data_1[3][0] = data_1[3][0] / 30 * 31 / data_1[3][3] 
#
#data_1[0][1] = data_1[0][1] / 31 * 31 / data_1[0][4] 
#data_1[1][1] = data_1[1][1] / 28 * 31 / data_1[1][4] 
#data_1[2][1] = data_1[2][1] / 31 * 31 / data_1[2][4] 
#data_1[3][1] = data_1[3][1] / 30 * 31 / data_1[3][4] 
#
#data_1[0][2] = data_1[0][2] / 31 * 31 / data_1[0][5] 
#data_1[1][2] = data_1[1][2] / 28 * 31 / data_1[1][5] 
#data_1[2][2] = data_1[2][2] / 31 * 31 / data_1[2][5] 
#data_1[3][2] = data_1[3][2] / 30 * 31 / data_1[3][5] 



print(data_1) 


    

result = []
import numpy as np  #多元线性回归，使用平均数得到账号数量值，带入回归方程，得到预测值
for j in range(3):
    x = np.matrix([data_1[0][3:7],data_1[1][3:7],data_1[2][3:7],data_1[3][3:7]],dtype=np.float64)
    y = np.matrix([data_1[0][j],data_1[1][j],data_1[2][j],data_1[3][j]],dtype = np.float64).T
    b=(x.T*x).I*x.T*y
    i = 0
    cb=[]
    while i<4:
        cb.append(b[i,0])
        i+=1
    temp_e=y-x*b
    mye=temp_e.sum()/temp_e.size
    e = np.matrix([mye,mye,mye]).T
    print('y=%f*x1+%f*x2+%f*x3+%f*x4+%f'%(cb[0],cb[1],cb[2],cb[3],mye))
    r = np.matrix([7,49,4223,31])#取平均值带入
    print(cb*r.T-mye) #得到预测值
    
    







